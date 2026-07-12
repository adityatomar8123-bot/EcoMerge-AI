"""
Excel Export Module for ESG Analysis Results

This module handles exporting ESG compliance analysis results to Excel format,
matching the structure of the MCG-Financials template.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from loguru import logger
import json


class ExcelExporter:
    """Export ESG analysis results to Excel format"""
    
    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize Excel Exporter
        
        Args:
            output_dir: Directory to save Excel files. Defaults to 'outputs/excel'
        """
        if output_dir is None:
            # Default to backend/outputs/excel directory
            self.output_dir = Path(__file__).parent.parent.parent / "outputs" / "excel"
        else:
            self.output_dir = Path(output_dir)
        
        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Excel exporter initialized with output directory: {self.output_dir}")
    
    def export_analysis_results(
        self,
        metric_analyses: List[Dict[str, Any]],
        industry: str,
        semi_industry: str,
        company_name: Optional[str] = None,
        report_id: Optional[str] = None
    ) -> Path:
        """
        Export analysis results to Excel file
        
        Args:
            metric_analyses: List of analyzed metrics with results
            industry: Main industry category
            semi_industry: Sub-industry for specific metrics
            company_name: Name of the company being analyzed
            report_id: Unique identifier for the report
            
        Returns:
            Path: Path to the generated Excel file
        """
        try:
            # Prepare data for Excel
            excel_data = []
            
            for metric in metric_analyses:
                row = {
                    "Metric": metric.get("metric_name", ""),
                    "Category": metric.get("category", ""),
                    "Unit": metric.get("unit", ""),
                    "Code": metric.get("metric_id", ""),
                    "Topic": metric.get("topic", ""),
                    "Type": metric.get("type", ""),
                    "Value": self._format_value(metric.get("value")),
                    "Page": self._format_page(metric.get("page")),
                    "Context": metric.get("context", ""),
                    "Disclosure Status": metric.get("disclosure_status", ""),
                    "LLM Analysis": metric.get("reasoning", "")
                }
                excel_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_company = self._sanitize_filename(company_name) if company_name else "Company"
            safe_industry = self._sanitize_filename(semi_industry)
            filename = f"{safe_company}_{safe_industry}_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main data
                sheet_name = self._truncate_sheet_name(semi_industry)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting
                self._apply_excel_formatting(worksheet, df)
                
                # Add metadata sheet
                metadata = {
                    "Analysis Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Company": company_name or "Unknown",
                    "Industry": industry,
                    "Sub-Industry": semi_industry,
                    "Report ID": report_id or "N/A",
                    "Total Metrics": len(metric_analyses),
                    "Fully Disclosed": sum(1 for m in metric_analyses if m.get("disclosure_status") == "fully_disclosed"),
                    "Partially Disclosed": sum(1 for m in metric_analyses if m.get("disclosure_status") == "partially_disclosed"),
                    "Not Disclosed": sum(1 for m in metric_analyses if m.get("disclosure_status") == "not_disclosed")
                }
                
                metadata_df = pd.DataFrame(list(metadata.items()), columns=["Field", "Value"])
                metadata_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Format metadata sheet
                metadata_worksheet = writer.sheets["Summary"]
                self._apply_metadata_formatting(metadata_worksheet, metadata_df)
            
            logger.info(f"Excel file successfully created: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            raise
    
    def _format_value(self, value: Any) -> Any:
        """Format value field for Excel"""
        if value is None or value == "null":
            return None
        if value == "not specific":
            return "Not Specific"
        return value
    
    def _format_page(self, page: Any) -> str:
        """Format page field for Excel"""
        if page is None or page == "null":
            return ""
        if isinstance(page, (list, tuple)):
            return ", ".join(str(p) for p in page)
        return str(page)
    
    def _sanitize_filename(self, name: str) -> str:
        """Sanitize string for use in filename"""
        if not name:
            return "Unknown"
        # Remove/replace invalid filename characters
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            name = name.replace(char, '_')
        # Limit length
        return name[:50]
    
    def _truncate_sheet_name(self, name: str) -> str:
        """Truncate sheet name to Excel's 31 character limit"""
        if len(name) > 31:
            return name[:28] + "..."
        return name
    
    def _apply_excel_formatting(self, worksheet, df):
        """Apply formatting to the main data worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Limit max width to 50 characters
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply text wrapping to Context and LLM Analysis columns
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        for row in range(2, len(df) + 2):
            # Context column (column 9)
            worksheet.cell(row=row, column=9).alignment = wrap_alignment
            # LLM Analysis column (column 11)
            worksheet.cell(row=row, column=11).alignment = wrap_alignment
        
        # Color code disclosure status
        status_colors = {
            "fully_disclosed": "C6EFCE",  # Light green
            "partially_disclosed": "FFEB9C",  # Light yellow
            "not_disclosed": "FFC7CE"  # Light red
        }
        
        status_col = 10  # Disclosure Status column
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=status_col)
            if cell.value in status_colors:
                cell.fill = PatternFill(start_color=status_colors[cell.value], 
                                       end_color=status_colors[cell.value], 
                                       fill_type="solid")
    
    def _apply_metadata_formatting(self, worksheet, df):
        """Apply formatting to the metadata worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True)
        for col in range(1, 3):
            worksheet.cell(row=1, column=col).font = header_font
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 40
        
        # Bold the field names
        field_font = Font(bold=True)
        for row in range(2, len(df) + 2):
            worksheet.cell(row=row, column=1).font = field_font
    
    def export_template(self, semi_industry: str) -> Path:
        """
        Export a blank template for a specific industry
        
        Args:
            semi_industry: Sub-industry name
            
        Returns:
            Path: Path to the generated template file
        """
        try:
            # Load metrics for the industry
            from .metric_processor import MetricProcessor
            from .models import ProcessingConfig
            
            config = ProcessingConfig()
            processor = MetricProcessor(config)
            metrics_collection = processor.load_sasb_metrics_by_industry(semi_industry)
            
            # Prepare template data
            template_data = []
            for metric in metrics_collection.metrics:
                row = {
                    "Metric": metric.metric_name,
                    "Category": metric.sasb_category if hasattr(metric, 'sasb_category') else "",
                    "Unit": metric.unit or "",
                    "Code": metric.metric_code,
                    "Topic": metric.sasb_topic if hasattr(metric, 'sasb_topic') else "",
                    "Type": metric.sasb_type if hasattr(metric, 'sasb_type') else "",
                    "Value": None,
                    "Page": None,
                    "Context": None
                }
                template_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(template_data)
            
            # Generate filename
            safe_industry = self._sanitize_filename(semi_industry)
            filename = f"Template_{safe_industry}.xlsx"
            filepath = self.output_dir / "templates" / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = self._truncate_sheet_name(semi_industry)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting
                worksheet = writer.sheets[sheet_name]
                self._apply_excel_formatting(worksheet, df)
            
            logger.info(f"Template Excel file created: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating template: {str(e)}")
            raise